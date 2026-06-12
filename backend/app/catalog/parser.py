"""Чтение прайс-файлов: xlsx (openpyxl) и csv (auto-кодировка, auto-разделитель)."""

import csv
import io
import zipfile
from dataclasses import dataclass

from openpyxl import load_workbook

Cell = str | float | int | None
Rows = list[list[Cell]]

HEADER_SCAN_LIMIT = 20


@dataclass
class ColumnInfo:
    index: int
    header: str
    samples: list[str]


def detect_header_row(rows: Rows, scan_limit: int = HEADER_SCAN_LIMIT) -> int:
    """Первая строка, где >=3 непустых ячеек и >=70% из них — текст. Иначе 0."""
    for i, row in enumerate(rows[:scan_limit]):
        filled = [c for c in row if c is not None and str(c).strip() != ""]
        if len(filled) < 3:
            continue
        text_cells = [c for c in filled if isinstance(c, str)]
        if len(text_cells) / len(filled) >= 0.7:
            return i
    return 0


def extract_columns(rows: Rows, header_row: int, sample_count: int = 3) -> list[ColumnInfo]:
    if header_row >= len(rows):
        return []
    header = rows[header_row]
    body = rows[header_row + 1 :]
    columns: list[ColumnInfo] = []
    for idx, cell in enumerate(header):
        title = str(cell).strip() if cell is not None else f"Колонка {idx + 1}"
        samples: list[str] = []
        for row in body:
            if len(samples) >= sample_count:
                break
            value = row[idx] if idx < len(row) else None
            if value is not None and str(value).strip() != "":
                samples.append(str(value))
        columns.append(ColumnInfo(index=idx, header=title, samples=samples))
    return columns


class UnsupportedFileError(Exception):
    pass


class CorruptFileError(Exception):
    pass


def load_tables(content: bytes, filename: str) -> dict[str, Rows]:
    """Файл -> {имя листа: строки}. Для csv единственный 'лист' с именем 'csv'."""
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _load_xlsx(content)
    if lower.endswith(".csv"):
        return {"csv": _load_csv(content)}
    raise UnsupportedFileError(filename)


def _load_xlsx(content: bytes) -> dict[str, Rows]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (zipfile.BadZipFile, KeyError, ValueError) as exc:
        raise CorruptFileError(str(exc)) from exc
    tables: dict[str, Rows] = {}
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        tables[ws.title] = rows
    wb.close()
    return tables


def _load_csv(content: bytes) -> Rows:
    text = _decode(content)
    delimiter = ";" if text.count(";") >= text.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(row) for row in reader]


def _decode(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")
