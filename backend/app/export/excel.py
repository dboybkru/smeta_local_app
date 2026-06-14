import io

from openpyxl import Workbook
from openpyxl.styles import Font


def _num(value) -> float:
    return float(value) if value is not None else 0.0


def render_xlsx(context: dict) -> bytes:
    """Лист: шапка → таблица разделов/позиций → итоги → место под подпись."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    ws["A1"] = "Смета / коммерческое предложение"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Объект: {context['object_name']}"

    proposal = context.get("proposal") or {}
    row = 4
    if context["level"] in ("full", "cover") and proposal.get("title"):
        ws.cell(row=row, column=1, value=proposal["title"]).font = Font(bold=True, size=12)
        row += 1
        if proposal.get("subtitle"):
            ws.cell(row=row, column=1, value=proposal["subtitle"])
            row += 1
        row += 1

    headers = ["Наименование", "Ед.", "Кол-во", "Материалы", "Работы", "Сумма"]
    for col, title in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=title).font = Font(bold=True)
    row += 1

    for section in context["sections"]:
        ws.cell(row=row, column=1, value=section["name"]).font = Font(bold=True)
        row += 1
        for ln in section["lines"]:
            mat = _num(ln["material_price"]) * _num(ln["qty"])
            work = _num(ln["work_price"]) * _num(ln["qty"])
            name = ln["name"]
            chars = ln.get("characteristics")
            if chars:
                name += " (" + "; ".join(f"{k}: {v}" for k, v in chars.items()) + ")"
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=ln["unit"])
            ws.cell(row=row, column=3, value=_num(ln["qty"]))
            ws.cell(row=row, column=4, value=mat)
            ws.cell(row=row, column=5, value=work)
            ws.cell(row=row, column=6, value=mat + work)
            row += 1
        st = section["totals"]
        ws.cell(row=row, column=5, value="Итого по разделу:").font = Font(bold=True)
        ws.cell(row=row, column=6, value=_num(st["total"])).font = Font(bold=True)
        row += 1

    totals = context["totals"]
    row += 1
    ws.cell(row=row, column=5, value="Без НДС:")
    ws.cell(row=row, column=6, value=_num(totals["subtotal"]))
    row += 1
    if context["vat_enabled"]:
        ws.cell(row=row, column=5, value="НДС:")
        ws.cell(row=row, column=6, value=_num(totals["vat"]))
        row += 1
    ws.cell(row=row, column=5, value="ВСЕГО:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=_num(totals["total"])).font = Font(bold=True)
    row += 3
    ws.cell(row=row, column=1, value="Подпись: ____________________   М.П.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
