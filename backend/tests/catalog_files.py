"""Генераторы файлов-фикстур: миниатюры реальных прайсов (Bolid, Optimus, работы)."""

import csv
import io

from openpyxl import Workbook


def make_bolid_xlsx() -> bytes:
    """Плоский прайс: заголовок в первой строке, розница + опт."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Болид"
    ws.append(["Название", "Описание", "Артикул", "Розничная_цена", "Оптовая_цена"])
    ws.append(["Сириус", "Прибор приемно-контрольный", "1-520-887", 36159.53, 33378.03])
    ws.append(["С2000-М", "Пульт контроля", "110-058-274", 12721.31, 11742.74])
    ws.append(["С2000-КДЛ", "Контроллер ДПЛС", "10-468-001", 4277.44, 3948.41])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_optimus_xlsx() -> bytes:
    """Многолистовой прайс: мусор перед заголовком, листы = категории."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "IP камеры"
    ws1.append(["Прайс-лист Optimus", None, None])
    ws1.append([None, None, None])
    ws1.append(["Модель", "Наименование", "Цена партнёра"])
    ws1.append(["IP-E012.1", "Видеокамера Optimus IP-E012.1", 3210.50])
    ws1.append(["IP-E014.0", "Видеокамера Optimus IP-E014.0", 5283.00])
    ws2 = wb.create_sheet("Сетевое оборудование")
    ws2.append(["Прайс-лист Optimus", None, None])
    ws2.append([None, None, None])
    ws2.append(["Модель", "Наименование", "Цена партнёра"])
    ws2.append(["U1E-8F", "Коммутатор Optimus U1E-8F", 6920.00])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_works_xlsx() -> bytes:
    """Прайс работ: имя/цена/ед.изм."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.append(["Наименование работы", "Цена руб.", "Ед. изм."])
    ws.append(["Монтаж камеры", 3500, "шт"])
    ws.append(["Прокладка кабеля", 150, "м"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_bolid_csv() -> bytes:
    """CSV-вариант плоского прайса (cp1251 — частая кодировка выгрузок)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["Название", "Артикул", "Розничная_цена", "Оптовая_цена"])
    writer.writerow(["Сириус", "1-520-887", "36159,53", "33378,03"])
    writer.writerow(["С2000-М", "110-058-274", "12721,31", "11742,74"])
    return buf.getvalue().encode("cp1251")
