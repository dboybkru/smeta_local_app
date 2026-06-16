import io

from openpyxl import load_workbook

from app.auth.models import User
from app.estimates.models import Estimate, EstimateBranch, EstimateLine, EstimateSection
from app.export import context as ctx
from app.export.excel import render_xlsx
from tests.orghelpers import get_or_create_org as _get_org


def _estimate(db_session):
    org = _get_org(db_session)
    u = User(email="u@x.ru", name="U", role="estimator", status="active", org_id=org.id)
    db_session.add(u)
    db_session.commit()
    est = Estimate(owner_id=u.id, org_id=org.id, object_name="Квартира")
    branch = EstimateBranch(name="Базовая")
    section = EstimateSection(name="Стены", markup_percent=10)
    section.lines.append(
        EstimateLine(name="Штукатурка", unit="м²", qty=10, work_price=300,
                     material_price=100, purchase_price_snapshot=80)
    )
    branch.sections.append(section)
    est.branches.append(branch)
    db_session.add(est)
    db_session.commit()
    db_session.refresh(est)
    return est


def test_public_context_strips_margin_and_purchase(db_session):
    est = _estimate(db_session)
    context = ctx.build_export_context(est, level="full", public=True)
    assert context["totals"]["margin"] is None
    assert context["totals"]["purchase"] is None
    for section in context["sections"]:
        assert section["totals"]["margin"] is None
        for line in section["lines"]:
            assert "purchase_price_snapshot" not in line


def test_private_context_keeps_margin(db_session):
    est = _estimate(db_session)
    context = ctx.build_export_context(est, level="full", public=False)
    assert context["totals"]["margin"] is not None
    # симметрично public-тесту: в приватном выводе маржа и закупка ПРИСУТСТВУЮТ на всех уровнях
    assert context["totals"]["purchase"] is not None
    for section in context["sections"]:
        assert section["totals"]["margin"] is not None
        for line in section["lines"]:
            assert "purchase_price_snapshot" in line


def test_export_includes_catalog_characteristics(db_session):
    from app.catalog.models import CatalogItem, Supplier
    org = _get_org(db_session)
    u = User(email="c@x.ru", name="U", role="estimator", status="active", org_id=org.id)
    db_session.add(u); db_session.commit()
    sup = Supplier(name="P", org_id=org.id); db_session.add(sup); db_session.commit()
    item = CatalogItem(supplier_id=sup.id, name="Камера", article="A", unit="шт",
                       kind="material", characteristics={"Разрешение": "2 Мп"}, org_id=org.id)
    db_session.add(item); db_session.commit()
    est = Estimate(owner_id=u.id, org_id=org.id, object_name="Объект")
    branch = EstimateBranch(name="Базовая")
    section = EstimateSection(name="Обор", markup_percent=0)
    section.lines.append(EstimateLine(item_id=item.id, name="Камера", unit="шт",
                                      qty=1, work_price=0, material_price=100))
    branch.sections.append(section); est.branches.append(branch)
    db_session.add(est); db_session.commit(); db_session.refresh(est)

    context = ctx.build_export_context(est, level="full", public=True, db=db_session)
    line = context["sections"][0]["lines"][0]
    assert line["characteristics"] == {"Разрешение": "2 Мп"}
    data = render_xlsx(context)
    wb = load_workbook(io.BytesIO(data))
    text = "\n".join(str(c.value) for row in wb.active.iter_rows() for c in row if c.value)
    assert "Разрешение" in text


def test_render_xlsx_is_valid_workbook(db_session):
    est = _estimate(db_session)
    context = ctx.build_export_context(est, level="estimate", public=False)
    data = render_xlsx(context)
    assert isinstance(data, bytes) and len(data) > 0
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "Квартира" in text
    assert "Штукатурка" in text
    assert "Стены" in text


def test_export_includes_client(db_session):
    from app.estimates.models import Client, Estimate, EstimateBranch
    org = _get_org(db_session)
    u = User(email="cl@x.ru", name="U", role="estimator", status="active", org_id=org.id)
    db_session.add(u); db_session.commit()
    cl = Client(name="ООО Ромашка", org_id=org.id, inn="7707083893", address="г Москва")
    db_session.add(cl); db_session.commit()
    est = Estimate(owner_id=u.id, org_id=org.id, object_name="Объект", client_id=cl.id)
    est.branches.append(EstimateBranch(name="Базовая"))
    db_session.add(est); db_session.commit(); db_session.refresh(est)
    context = ctx.build_export_context(est, level="full", public=False, db=db_session)
    assert context["client"]["name"] == "ООО Ромашка"
    assert context["client"]["inn"] == "7707083893"
    data = render_xlsx(context)
    wb = load_workbook(io.BytesIO(data))
    text = "\n".join(str(c.value) for row in wb.active.iter_rows() for c in row if c.value)
    assert "Ромашка" in text
